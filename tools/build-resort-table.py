#!/usr/bin/env python3
"""Build docs/resort-table.xlsx — the q25 fact/ratings table for Pingwin's approval.
Source: config/resort-profiles.json + config/departures.json + data/camps.json."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(__file__), '..')
prof = json.load(open(os.path.join(ROOT, 'config/resort-profiles.json'), encoding='utf-8'))
dep = json.load(open(os.path.join(ROOT, 'config/departures.json'), encoding='utf-8'))
camps = json.load(open(os.path.join(ROOT, 'data/camps.json'), encoding='utf-8'))
camp_resorts = set(camps['resorts'])

COUNTRY_HE = {'austria': 'אוסטריה', 'bulgaria': 'בולגריה', 'andorra': 'אנדורה', 'france': 'צרפת'}
# draft judgements → a 1–5 scale (5 = the most). Same scale Pingwin fills in.
NIGHT_SCORE = {'party': 5, 'lively': 4, 'moderate': 3, 'quiet': 2}
FAM_SCORE = {'high': 5, 'medium': 3, 'low': 1}

def yn(v):
    return '' if v is None else ('כן' if v else 'לא')

ARIAL = 'Arial'
F = lambda **k: Font(name=ARIAL, **k)
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
SUB_FILL = PatternFill('solid', fgColor='D9E1F2')
EDIT_FILL = PatternFill('solid', fgColor='FFF2CC')   # yellow: Pingwin fills
DRAFT_FILL = PatternFill('solid', fgColor='FCE4D6')  # orange: draft judgement, please confirm
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = 'אתרים'
ws.sheet_view.rightToLeft = True

# (header, width, kind, group) kind: fact | draft | edit
COLS = [
    ('אתר', 18, 'fact', 'זיהוי'), ('מדינה', 10, 'fact', 'זיהוי'), ('שם באנגלית', 20, 'fact', 'זיהוי'),
    ('גובה הכפר (מ׳)', 10, 'fact', 'ההר'), ('גובה מקסימלי (מ׳)', 10, 'fact', 'ההר'),
    ('ק"מ מסלולים באתר', 10, 'fact', 'ההר'), ('ק"מ באזור המקושר', 10, 'fact', 'ההר'), ('שם האזור המקושר', 20, 'fact', 'ההר'),
    ('% מסלולים קלים', 9, 'fact', 'ההר'), ('% מסלולים קשים', 9, 'fact', 'ההר'),
    ('קרחון', 8, 'fact', 'ההר'), ('סנואו-פארק', 9, 'fact', 'ההר'), ('סקי לילה', 8, 'fact', 'ההר'),
    ('אזור מתחילים ליד הכפר', 11, 'fact', 'נוחות'), ('Ski-in / Ski-out', 10, 'fact', 'נוחות'),
    ('שדה תעופה', 10, 'fact', 'לוגיסטיקה'), ('ק"מ משדה התעופה', 10, 'fact', 'לוגיסטיקה'),
    ('נציג פינגווין', 16, 'fact', 'לוגיסטיקה'), ('קייטנה בעברית', 9, 'fact', 'לוגיסטיקה'), ('עיר קרובה', 18, 'fact', 'לוגיסטיקה'),
    ('אפרה-סקי 1–5\n(הצעה שלנו)', 11, 'draft', 'הצעה שלנו — לאשר/לתקן'), ('משפחות 1–5\n(הצעה שלנו)', 11, 'draft', 'הצעה שלנו — לאשר/לתקן'),
    ('מתחילים 1–5', 9, 'edit', 'פינגווין ממלאת'), ('משפחות עם ילדים 1–5', 10, 'edit', 'פינגווין ממלאת'), ('אפרה-סקי / צעירים 1–5', 11, 'edit', 'פינגווין ממלאת'),
    ('גולשים מנוסים 1–5', 9, 'edit', 'פינגווין ממלאת'), ('רמת מחיר\n(משתלם / בינוני / פרימיום)', 13, 'edit', 'פינגווין ממלאת'),
    ('להמליץ? (כן/לא)', 9, 'edit', 'פינגווין ממלאת'), ('הערות פינגווין', 30, 'edit', 'פינגווין ממלאת'), ('מאושר ✓', 8, 'edit', 'פינגווין ממלאת'),
    ('מקור', 40, 'fact', 'מקור'),
]

ws['A1'] = 'טבלת אתרים — בסיס להמלצה מנומקת (שאלה 25)'
ws['A1'].font = F(bold=True, size=14)
ws['A2'] = 'לבן = עובדה מהאתר הרשמי.  כתום = הצעה שלנו (1–5), לאשר או לתקן.  צהוב = פינגווין ממלאת.  כל הסולמות מוסברים בגיליון "מקרא".  הבוט ישתמש רק בשורות עם ✓ ב"מאושר".'
ws['A2'].font = F(italic=True, size=10)
ws.merge_cells('A1:J1'); ws.merge_cells('A2:T2')

GR, HR = 4, 5
GROUP_FILL = {'זיהוי': '1F3864', 'ההר': '2F5597', 'נוחות': '2F5597', 'לוגיסטיקה': '2F5597',
              'הצעה שלנו — לאשר/לתקן': 'C55A11', 'פינגווין ממלאת': 'BF8F00', 'מקור': '7F7F7F'}
# group band row
start = 1
for i in range(1, len(COLS) + 2):
    g_prev = COLS[start - 1][3]
    if i > len(COLS) or COLS[i - 1][3] != g_prev:
        ws.merge_cells(start_row=GR, start_column=start, end_row=GR, end_column=i - 1)
        c = ws.cell(row=GR, column=start, value=g_prev)
        c.font = F(bold=True, color='FFFFFF', size=11); c.fill = PatternFill('solid', fgColor=GROUP_FILL[g_prev]); c.alignment = CENTER
        for j in range(start, i): ws.cell(row=GR, column=j).border = BORDER
        start = i
for i, (h, w, kind, g) in enumerate(COLS, 1):
    c = ws.cell(row=HR, column=i, value=h)
    c.font = F(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=GROUP_FILL[g]); c.alignment = CENTER; c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[HR].height = 52

order = sorted(prof['resorts'].items(), key=lambda kv: (['austria', 'bulgaria', 'andorra', 'france'].index(kv[1]['country']), kv[0]))
r = HR + 1
for name, p in order:
    t = dep['transfer_km'].get(name, {})
    rep = dep['reps'].get(name, {})
    rep_he = 'באתר' if rep.get('on_site') else (f"מ-{prof['resorts'].get(rep.get('served_from'), {}).get('he', rep.get('served_from'))}" if rep.get('served_from') else '')
    if rep.get('note_he'):
        rep_he += f" ({rep['note_he']})"
    km = t.get('km')
    if t.get('alt'):
        km_txt = f"{km} (או {t['alt']['km']} מ{t['alt']['airport']})"
    else:
        km_txt = km
    row = [p['he'], COUNTRY_HE[p['country']], name, p['village_m'], p['top_m'], p['piste_km'], p['linked_km'], p['linked_name'],
           p['easy_pct'], p['hard_pct'], yn(p['glacier']), yn(p['snow_park']), yn(p['night_skiing']),
           yn(p['beginner_near_village']), yn(p['ski_in_out']), t.get('airport'), km_txt, rep_he,
           'כן' if name in camp_resorts else 'לא', p['city_he'],
           NIGHT_SCORE.get(p['nightlife']), FAM_SCORE.get(p['family']),
           None, None, None, None, None, None, None, None, p['source']]
    for i, v in enumerate(row, 1):
        c = ws.cell(row=r, column=i, value=v)
        kind = COLS[i - 1][2]
        c.font = F(size=10); c.border = BORDER
        c.alignment = RIGHT if i in (1, 3, 8, 18, 20, 29, 31) else CENTER
        if kind == 'draft' and isinstance(v, int): c.font = F(size=10, bold=True)
        if kind == 'draft': c.fill = DRAFT_FILL
        elif kind == 'edit': c.fill = EDIT_FILL
        if v is None and kind == 'fact':
            c.value = '?'; c.font = F(size=10, color='C00000')
    r += 1
LAST = r - 1
ws.freeze_panes = ws.cell(row=HR + 1, column=2)
ws.auto_filter.ref = f"A{HR}:{get_column_letter(len(COLS))}{LAST}"

# validations on the editable columns
def col(h):
    return get_column_letter([c[0] for c in COLS].index(h) + 1)
dv15 = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
dvyn = DataValidation(type='list', formula1='"כן,לא"', allow_blank=True)
dvbud = DataValidation(type='list', formula1='"משתלם,בינוני,פרימיום"', allow_blank=True)
dvok = DataValidation(type='list', formula1='"✓"', allow_blank=True)
for dv in (dv15, dvyn, dvbud, dvok): ws.add_data_validation(dv)
for h in ('מתחילים 1–5', 'משפחות עם ילדים 1–5', 'אפרה-סקי / צעירים 1–5', 'גולשים מנוסים 1–5', 'אפרה-סקי 1–5\n(הצעה שלנו)', 'משפחות 1–5\n(הצעה שלנו)'):
    dv15.add(f"{col(h)}{HR+1}:{col(h)}{LAST}")
dvyn.add(f"{col('להמליץ? (כן/לא)')}{HR+1}:{col('להמליץ? (כן/לא)')}{LAST}")
PRICE_H = 'רמת מחיר\n(משתלם / בינוני / פרימיום)'
dvbud.add(f"{col(PRICE_H)}{HR+1}:{col(PRICE_H)}{LAST}")
dvok.add(f"{col('מאושר ✓')}{HR+1}:{col('מאושר ✓')}{LAST}")

# comments on low-confidence facts
notes = {
    ('Flaine Grand Massif', 'ק"מ משדה התעופה'): 'לא ודאי — 70 ק"מ מז\'נבה נראה נמוך; נא לאמת.',
    ('Montgenevre', 'ק"מ משדה התעופה'): 'לא ודאי — 215 ק"מ מליון; נא לאמת (מטורינו ~90 ק"מ).',
}
for (name, h), txt in notes.items():
    rr = HR + 1 + [n for n, _ in order].index(name)
    ws[f"{col(h)}{rr}"].comment = Comment(txt, 'Ai-Assistant')
    ws[f"{col(h)}{rr}"].fill = DRAFT_FILL

# summary line
ws.cell(row=LAST + 2, column=1, value='מאושרים:').font = F(bold=True)
ws.cell(row=LAST + 2, column=2, value=f'=COUNTIF({col("מאושר ✓")}{HR+1}:{col("מאושר ✓")}{LAST},"✓")&" מתוך "&COUNTA(A{HR+1}:A{LAST})').font = F(bold=True)
ws.cell(row=LAST + 3, column=1, value='"?" = נתון שלא נמצא במקור הרשמי. נא להשלים או להשאיר ריק — הבוט לא ימציא.').font = F(italic=True, size=9)

# Legend sheet
lg = wb.create_sheet('מקרא')
lg.sheet_view.rightToLeft = True
lg.column_dimensions['A'].width = 30; lg.column_dimensions['B'].width = 100
def lrow(r, a, b, bold=False, fill=None, size=10):
    ca = lg.cell(row=r, column=1, value=a); cb = lg.cell(row=r, column=2, value=b)
    ca.font = F(bold=True, size=size); cb.font = F(size=size, bold=bold); cb.alignment = RIGHT; ca.alignment = RIGHT
    if fill: ca.fill = fill
    lg.row_dimensions[r].height = 18 if len(b) < 90 else 34
r = 1
lrow(r, 'מה הטבלה הזאת', 'עובדות + דירוגים לכל אתר. אחרי אישור, הבוט ממליץ מנומק: "טיניי מתאים לכם כי: אתר גבוה עם קרחון, קייטנה בעברית, נציג באתר". בלי אישור — אין המלצות.', size=11); r += 2
lrow(r, 'צבעים', '', bold=True); r += 1
lrow(r, 'לבן', 'עובדה מהאתר הרשמי (קישור בעמודה "מקור"). מותר לתקן. "?" = לא נמצא — להשלים או להשאיר ריק, הבוט לא ימציא.'); r += 1
lrow(r, 'כתום', 'הצעה שלנו בסולם 1–5, מתוך המיצוב של האתר עצמו. לא מחייב — לאשר או לשנות את המספר.', fill=DRAFT_FILL); r += 1
lrow(r, 'צהוב', 'פינגווין ממלאת: דירוג 1–5 לכל קהל, רמת מחיר, להמליץ?, הערות, ו-✓ ב"מאושר".', fill=EDIT_FILL); r += 2
lrow(r, 'הסולם 1–5 — כלל', '5 = האתר הכי מתאים לקהל הזה אצלנו, 1 = לא מתאים. הבוט ימליץ לקהל רק על אתרים עם 4–5, ויסביר למה.', bold=True); r += 2
lrow(r, 'אפרה-סקי / צעירים', '', bold=True); r += 1
for k, v in [('5', 'בירת אפרה-סקי: ברים על המסלול מהצהריים, מועדונים, מסיבות עד הלילה (אישגל, ואל טורנס, פאס דה לה קאסה, לה דוז אלפ)'),
             ('4', 'תוסס: הרבה ברים ומסעדות, מוזיקה חיה, מועדון או שניים (בנסקו, בורובץ, טיניי, סולדו)'),
             ('3', 'יש חיים בערב: כמה ברים ומסעדות, בלי סצנת מסיבות (אבוריאז, לה מנואר, לה ארק)'),
             ('2', 'שקט: מסעדות, ערב במלון, מתאים למי שבא לישון מוקדם ולגלוש (פליין, מונז\'נבר, עוז)'),
             ('1', 'כמעט כלום בערב')]:
    lrow(r, k, v); r += 1
r += 1
lrow(r, 'משפחות עם ילדים', '', bold=True); r += 1
for k, v in [('5', 'בנוי למשפחות: מתחילים ליד הכפר, בלי מכוניות / ski-in-out, קייטנה בעברית או גן שלג, מלונות משפחתיים'),
             ('4', 'נוח למשפחות, בלי אחד מהתנאים למעלה'),
             ('3', 'אפשרי: מסתדרים, אבל צריך אוטובוס למסלול או שהכפר רועש בלילה'),
             ('2', 'פחות מתאים: תלול, רועש, או לוגיסטיקה מסובכת עם ילדים'),
             ('1', 'לא ממליצים למשפחות')]:
    lrow(r, k, v); r += 1
r += 1
lrow(r, 'מתחילים', '', bold=True); r += 1
for k, v in [('5', 'אזור מתחילים רחב ליד הכפר, מסלולים ירוקים/כחולים ארוכים, ski school חזק באנגלית'),
             ('3', 'יש איפה ללמוד, אבל המסלולים הקלים מוגבלים או רחוקים מהכפר'),
             ('1', 'אתר לגולשים מנוסים — מתחיל יסבול')]:
    lrow(r, k, v); r += 1
r += 1
lrow(r, 'גולשים מנוסים', '', bold=True); r += 1
for k, v in [('5', 'אזור ענק (300+ ק"מ מקושרים), גובה, קרחון, מסלולים שחורים ואוף-פיסט'),
             ('3', 'אתר בינוני — כיף לשבוע, לא מאתגר'),
             ('1', 'קטן / נמוך — מנוסה ישתעמם אחרי יומיים')]:
    lrow(r, k, v); r += 1
r += 1
lrow(r, 'רמת מחיר', 'מיצוב יחסי בין האתרים בלבד: משתלם / בינוני / פרימיום. הבוט לעולם לא יציג מספר, רק את המילה.'); r += 1
lrow(r, 'להמליץ?', '"לא" = הבוט יענה על שאלות על האתר, אבל לא יציע אותו מיוזמתו.'); r += 1
lrow(r, 'מאושר ✓', 'רק שורות עם ✓ נכנסות למנוע ההמלצות. אפשר לאשר בהדרגה, שורה-שורה.'); r += 1
lrow(r, 'שורת דוגמה', 'שורת טיניי מולאה בצהוב כדוגמה לפורמט (אפור נטוי) — נא לדרוס.'); r += 1
lrow(r, 'מה לא כאן בכוונה', 'מחירים, שמות מלונות, זמני נסיעה — לפי הכללים האדומים.')

# example row: Tignes
rr = HR + 1 + [n for n, _ in order].index('Tignes')
ex = {'מתחילים 1–5': 3, 'משפחות עם ילדים 1–5': 5, 'אפרה-סקי / צעירים 1–5': 4, 'גולשים מנוסים 1–5': 5,
      PRICE_H: 'פרימיום', 'להמליץ? (כן/לא)': 'כן', 'הערות פינגווין': 'דוגמה בלבד — נא לעדכן'}
for h, v in ex.items():
    c = ws[f"{col(h)}{rr}"]; c.value = v; c.font = F(size=10, italic=True, color='7F7F7F')

out = os.path.join(ROOT, 'docs/resort-table.xlsx')
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(out)
