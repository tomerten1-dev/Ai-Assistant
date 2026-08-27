#!/usr/bin/env python3
"""Read docs/resort-table.xlsx (Pingwin's approved copy) back into config/resort-profiles.json.
Usage: python3 tools/ingest-resort-table.py [--all]   (--all = treat every row as approved, for the demo)"""
import json, os, sys, datetime
from openpyxl import load_workbook

ROOT = os.path.join(os.path.dirname(__file__), '..')
P = os.path.join(ROOT, 'config/resort-profiles.json')
prof = json.load(open(P, encoding='utf-8'))
ws = load_workbook(os.path.join(ROOT, 'docs/resort-table.xlsx'), data_only=True)['אתרים']
HR = 5
hdr = {ws.cell(row=HR, column=c).value: c for c in range(1, ws.max_column + 1)}
ALL = '--all' in sys.argv
PRICE = {'משתלם': 'budget', 'בינוני': 'mid', 'פרימיום': 'premium'}

def yn(v):
    if v in (None, '', '?'): return None
    s = str(v).strip()
    if s.startswith('כן'): return True
    if s.startswith('לא'): return False
    if s.startswith('חלקי'): return 'partial'
    return None

def num(v):
    try: return int(v) if v not in (None, '', '?') else None
    except (TypeError, ValueError): return None

def cell(r, h): return ws.cell(row=r, column=hdr[h]).value

approved = 0
for r in range(HR + 1, HR + 1 + len(prof['resorts'])):
    name = cell(r, 'שם באנגלית')
    if name not in prof['resorts']: continue
    p = prof['resorts'][name]
    ok = ALL or (cell(r, 'מאושר ✓') == '✓')
    p['village_m'] = num(cell(r, 'גובה הכפר (מ׳)')); p['top_m'] = num(cell(r, 'גובה מקסימלי (מ׳)'))
    p['piste_km'] = num(cell(r, 'ק"מ מסלולים באתר')); p['linked_km'] = num(cell(r, 'ק"מ באזור המקושר'))
    p['easy_pct'] = num(cell(r, '% מסלולים קלים')); p['hard_pct'] = num(cell(r, '% מסלולים קשים'))
    p['glacier'] = yn(cell(r, 'קרחון')); p['snow_park'] = yn(cell(r, 'סנואו-פארק')); p['night_skiing'] = yn(cell(r, 'סקי לילה'))
    p['beginner_near_village'] = yn(cell(r, 'אזור מתחילים ליד הכפר')); p['ski_in_out'] = yn(cell(r, 'Ski-in / Ski-out'))
    p['ratings'] = {
        'beginners': num(cell(r, 'מתחילים 1–5')), 'families': num(cell(r, 'משפחות עם ילדים 1–5')),
        'apres': num(cell(r, 'אפרה-סקי / צעירים 1–5')), 'experts': num(cell(r, 'גולשים מנוסים 1–5')),
    }
    p['price_level'] = PRICE.get(str(cell(r, 'רמת מחיר\n(משתלם / בינוני / פרימיום)') or '').strip())
    p['recommend'] = (cell(r, 'להמליץ? (כן/לא)') or 'כן') == 'כן'
    notes = str(cell(r, 'הערות פינגווין') or '')
    p['reason_he'] = notes.split('\n')[0].strip()
    p['approved'] = ok
    for k in ('nightlife', 'family'): p.pop(k, None)   # replaced by ratings
    approved += ok

prof['status'] = 'approved_demo' if ALL else 'partial'
prof['approved_by'] = 'תומר — אישור זמני לדמו, לשיפור בהמשך' if ALL else prof.get('approved_by')
prof['approved_at'] = datetime.date.today().isoformat()
prof['_comment'] = ('פרופיל עובדתי + דירוגים 1–5 לכל אתר (מתחילים/משפחות/אפרה-סקי/מנוסים), רמת מחיר יחסית, ונימוק. '
                    'הבסיס ל"המלצה מנומקת" (שאלה 25). הדירוגים והתיקונים מתוך מחקר אינטרנט, אושרו זמנית לדמו ע"י תומר (docs/resort-table.xlsx). '
                    'עריכה: בטבלה → python3 tools/ingest-resort-table.py. הבוט משתמש רק ב-approved=true.')
json.dump(prof, open(P, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
print(f'{approved}/{len(prof["resorts"])} approved → {P}')
